import pyautogui
import openpyxl
from time import sleep

from openpyxl import Workbook
from openpyxl import load_workbook

### Declarations/Initializations
routes        = Workbook()                                                       # initializes new workbook
dest_filename = 'optroutes.xlsx'                                                 # filename for new excel doc
stormog       = load_workbook(filename = 'Storm Services.xlsx')                  # open the storm services master

scurrent_ws: object = stormog.get_sheet_by_name('DIS & GDT Osmose Pole Audits')  # make the audit sheet the current sheet
ocurrent_ws         = routes.create_sheet('Pole Audit Routes', 0)                # create and make new sheet the current sheet

Route_Order = []
Points_List = []
Primary_P   = starting_loc
Start_P     = node_searching_from
Next_P      = node_closest_to_start

i           = 0
rowNum      = 1
n           = 1
StructNums  = []
FeederNames = scurrent_ws.cell(row=rowNum, column=2).value                       # for future for loop
Feeder      = 'Rankin'                                                           # changes depending on which feeder is being concerned

GTV            = pyautogui.locateOnScreen('GTV.png')
GTVx, GTVy     = pyautogui.center(GTV)

# Copy the data needed from storm services master and write to new excel doc
for rowNum in range(2, scurrent_ws.max_row):
    FeederNames = scurrent_ws.cell(row=rowNum, column=2).value
    if FeederNames == Feeder:
        StructNum = scurrent_ws.cell(row = rowNum, column = 11).value
        StructNums.append(StructNum)

 # test code for above block
 print(StructNums)
 print(len(StructNums))

for i in range(len(StructNums)):
    ocurrent_ws.cell(row=n, column=1).value = StructNums[i]
    i += 1
    n += 1





routes.save(filename = dest_filename)

#open GT Viewer
pyautogui.moveTo(GTVx, GTVy)
pyautogui.doubleClick(GTVx, GTVy)
print('Opening program...')
sleep(30)

###select I accept button
GTV_A          = pyautogui.locateOnScreen('Accept.png')
print(GTV_A)
GTV_Ax, GTV_Ay = pyautogui.center(GTV_A)
print(GTV_Ax, GTV_Ay)
pyautogui.click(GTV_Ax, GTV_Ay)
sleep(2)

#select query, then by pole type
GTV_Q          = pyautogui.locateOnScreen('Query.png')
GTV_Qx, GTV_Qy = pyautogui.center(GTV_Q)
pyautogui.click(GTV_Qx, GTV_Qy)
for a in range(1, 15[ 1]):
    pyautogui.press('down')
